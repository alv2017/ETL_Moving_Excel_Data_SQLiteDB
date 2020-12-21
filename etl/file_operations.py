import csv
import os
from openpyxl import load_workbook
from settings import DATA_DIR, READ_LOG_FILE
from etl.logging_errors import copy_errors_to_error_log

def log_missing_price_error(logger, data_tuple, ref_file):
    ref_file_basename = os.path.basename(ref_file)
    msg = "{0}: The price is missing or not valid: {1}.".format(ref_file_basename, str(data_tuple))
    logger.error(msg)
    
def log_missing_data_summary(logger, ref_file, missing_data_entries, log_errors_to_file=True):
    if missing_data_entries > 0:
        ref_file_basename = os.path.basename(ref_file)
        ref_file_name, ref_file_ext = ref_file_basename.split(".")
        msg = "There is missing price data in the data file: {0}\n".format(ref_file_basename)
        msg += "!!! WARNING: Number of missing values: {0} !!!\n".format(missing_data_entries)
        # File extension
        if log_errors_to_file:
            readlog_name = "read_error_" + ref_file_name + "_" + ref_file_ext + ".log"
            msg += "  For details check file's read log: {0}.".format(readlog_name)
        msg += "\n"
        logger.info(msg)
        
def process_data_row(data_row, region_list):
    nregions = len(region_list)
    # price time
    date = data_row[0].split("-")
    day = date[0]
    month = date[1]
    year = date[2]
    hour = data_row[1][:2]    
    dt = str(year) + "-" + str(month) + "-" + str(day) + " " + str(hour) + ":" + "00"
    
    # Regions and Price
    for i in range(nregions):
        price_value = data_row[i+2]
        
        try: 
            price = str( "%.2f" % float(price_value) )
        except:
            price = ''
            
        data_tuple = (dt, region_list[i], price)
        yield data_tuple

def read_csv_file_data(csv_file, app_logger, read_logger, 
                       log_errors_to_file=True, skip_n_rows = 2):
    """ The function reads data from the pre-formatted csv file and logs data errors
        to the corresponding log file.
        :params csv_file: csv file from which we are going to read data
        :type csv_file: str
        :params skip_n_rows: the number of rows to skip, the default value is 2
        :type skip_n_rows: int
        :return: data tuples generator. Data tuples are of the form (price_time, region, price) 
    """
    # log info
    msg = "Reading data from:\n  {0}".format(csv_file)
    app_logger.info(msg)
    
    # missing data counter
    missing_data_entries = 0

    with open(csv_file, newline='') as csvf:
        datareader = csv.reader(csvf, delimiter = ',', quotechar = '"')
        
        # Skipping n rows
        for i in range(skip_n_rows):
            next(datareader)
            
        # Regions
        regions = next(datareader)[2:]
            
        # Data processing
        for row in datareader:
            for processed_row in process_data_row(row, regions):
                if processed_row[2] != '':
                    yield processed_row
                else:
                    missing_data_entries += 1
                    log_missing_price_error(read_logger, processed_row, csv_file)
                    
    # Log info on missing price data (if any)
    if missing_data_entries > 0:
        if log_errors_to_file==True:
            copy_errors_to_error_log("read_error", csv_file, READ_LOG_FILE)
        log_missing_data_summary(app_logger, csv_file, missing_data_entries, 
                                 log_errors_to_file=log_errors_to_file)

        
def read_excel_file_data(excel_file, app_logger, read_logger, log_errors_to_file=True, wsname = None):
    """ The function reads data from the pre-formatted excel file and logs data errors
        to the corresponding log file.
        :params csv_file: csv file from which we are going to read data
        :type csv_file: str
        :params skip_n_rows: the number of rows to skip, the default value is 2
        :type skip_n_rows: int
        :return: data tuples generator. Data tuples are of the form (price_time, region, price) 
    """
    
    # log info
    msg = "Reading data from:\n  {0}".format(excel_file)
    app_logger.info(msg)
    
    # missing data counter
    missing_data_entries = 0
    
    if wsname is None:
        wsname = os.path.basename(excel_file).split(".")[0]
    
    wb = load_workbook(excel_file)
    ws = wb[wsname]
    
    # regions
    regions = []
    for cell in ws[3][2:]:
        regions.append(cell.value)

    # data records
    for row in ws.iter_rows(min_row=4, min_col=1, values_only=True):
        for processed_row in process_data_row(row, regions):
            if processed_row[2] != '':
                yield processed_row
            else:
                missing_data_entries += 1
                log_missing_price_error(read_logger, processed_row, excel_file)
        
    # Log info on missing price data (if any)
    if missing_data_entries > 0:
        if log_errors_to_file==True:
            copy_errors_to_error_log("read_error", excel_file, READ_LOG_FILE)
        log_missing_data_summary(app_logger, excel_file, missing_data_entries, 
                                 log_errors_to_file=log_errors_to_file)
        
if __name__ == "__main__":
    from logger.testing_logger import logger as testlogger
    from logger.app_logger import logger as applogger
    from logger.read_logger import logger as readlogger

    
    
    region_list = ['SYS', 'SE1', 'SE2', 'SE3', 'SE4', 'FI', 'DK1', 'DK2', 'Oslo', 
               'Kr.sand', 'Bergen', 'Molde', 'Tr.heim', 'Tromsø', 'EE', 'LV', 'LT']

    data_row_1 = ["01-01-2018", "14\xa0-\xa015", "26.03", "25.81", "25.81", "25.81", "25.81",
                "25.81", "9.15", "25.81", "26.19", "26.19", "26.19", "26.19", "26.19",
                "26.19", "", "", "25.81"]

    data_row_2 = ['25-03-2018', '02\xa0-\xa003', '', '', '', '', 
                '', '', '', '', '', '', '', '', '', '', '', '', '']
    
    import inspect
    """
    data = process_data_row(data_row_1, region_list)
    if inspect.isgenerator(data):
        print("TRUE")
    else:
        print("FALSE")
    print(data)
    
    """                
    excel_file_name = "elspot-prices_2018_hourly_eur.xlsx"
    csv_file_name = "elspot-prices_2018_hourly_eur.csv"
    
    excel_file = os.path.join(DATA_DIR, excel_file_name)
    csv_file = os.path.join(DATA_DIR, csv_file_name)
    
    excel_data = read_excel_file_data(excel_file, applogger, readlogger, log_errors_to_file=True)
    csv_data = read_csv_file_data(csv_file, applogger, readlogger, log_errors_to_file=True)
    
    nExcelRows = 0
    for row in excel_data:
        nExcelRows += 1
        
    nCSVRows = 0
    for row in csv_data:
        nCSVRows += 1
        
    print("Number of rows in CSV file: ", nCSVRows)
    print("Number of rows in Excel file: ", nExcelRows)
    
    